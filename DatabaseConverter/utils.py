import re
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Any, Optional
import json

def validate_company_number(company_number: str) -> bool:
    """Validate UK company number format"""
    if not company_number:
        return False
    
    # Remove any spaces or special characters
    clean_number = re.sub(r'[^A-Z0-9]', '', company_number.upper())
    
    # UK company numbers are typically 8 digits, sometimes with 2-letter prefix
    # Examples: 12345678, SC123456, NI123456, OC123456
    patterns = [
        r'^\d{8}$',  # 8 digits
        r'^[A-Z]{2}\d{6}$',  # 2 letters + 6 digits
        r'^[A-Z]{1}\d{7}$',  # 1 letter + 7 digits
    ]
    
    return any(re.match(pattern, clean_number) for pattern in patterns)

def format_company_data(company_data: Dict) -> Dict:
    """Format company data for display"""
    formatted = {}
    
    # Basic fields
    formatted['Company Name'] = company_data.get('company_name', 'N/A')
    formatted['Company Number'] = company_data.get('company_number', 'N/A')
    formatted['Status'] = company_data.get('company_status', 'N/A').title()
    formatted['Type'] = company_data.get('company_type', 'N/A')
    formatted['Jurisdiction'] = company_data.get('jurisdiction', 'N/A')
    
    # Dates
    if 'date_of_creation' in company_data:
        formatted['Incorporated'] = company_data['date_of_creation']
    
    # Address formatting
    if 'registered_office_address' in company_data:
        address_parts = []
        addr_data = company_data['registered_office_address']
        
        for key in ['address_line_1', 'address_line_2', 'locality', 'region', 'postal_code', 'country']:
            if key in addr_data and addr_data[key]:
                address_parts.append(addr_data[key])
        
        formatted['Address'] = ', '.join(address_parts)
    
    # SIC codes
    if 'sic_codes' in company_data and company_data['sic_codes']:
        formatted['SIC Codes'] = ', '.join(company_data['sic_codes'])
    
    return formatted

def export_to_excel(companies_df: pd.DataFrame, include_enrichment: bool = True) -> bytes:
    """Export companies data to Excel format"""
    
    # Create workbook
    wb = Workbook()
    
    # Companies sheet
    ws_companies = wb.active
    ws_companies.title = "Companies"
    
    # Prepare companies data
    export_columns = [
        'company_name', 'company_number', 'company_status', 'company_type',
        'jurisdiction', 'date_of_creation', 'address', 'sic_codes', 'created_at'
    ]
    
    companies_export = companies_df[export_columns].copy()
    
    # Add headers and data
    for r in dataframe_to_rows(companies_export, index=False, header=True):
        ws_companies.append(r)
    
    # Style the headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws_companies[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws_companies.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws_companies.column_dimensions[column_letter].width = adjusted_width
    
    # Enrichment data sheet (if requested)
    if include_enrichment and 'enrichment_data' in companies_df.columns:
        ws_enrichment = wb.create_sheet("Enrichment Data")
        
        enrichment_data = []
        for idx, row in companies_df.iterrows():
            if pd.notna(row['enrichment_data']):
                try:
                    if isinstance(row['enrichment_data'], str):
                        enrich_data = json.loads(row['enrichment_data'])
                    else:
                        enrich_data = row['enrichment_data']
                    
                    for provider, data in enrich_data.items():
                        if data:
                            flat_data = flatten_dict(data, f"{provider}_")
                            flat_data.update({
                                'company_name': row['company_name'],
                                'company_number': row['company_number'],
                                'provider': provider
                            })
                            enrichment_data.append(flat_data)
                except:
                    continue
        
        if enrichment_data:
            enrichment_df = pd.DataFrame(enrichment_data)
            
            for r in dataframe_to_rows(enrichment_df, index=False, header=True):
                ws_enrichment.append(r)
            
            # Style headers
            for cell in ws_enrichment[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws_enrichment.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws_enrichment.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def flatten_dict(d: Dict, parent_key: str = '', sep: str = '_') -> Dict:
    """Flatten nested dictionary"""
    items = []
    
    for k, v in d.items():
        new_key = f"{parent_key}{k}" if parent_key else k
        
        if isinstance(v, dict):
            items.extend(flatten_dict(v, f"{new_key}{sep}", sep=sep).items())
        elif isinstance(v, list):
            items.append((new_key, ', '.join(str(item) for item in v)))
        else:
            items.append((new_key, v))
    
    return dict(items)

def clean_company_name(company_name: str) -> str:
    """Clean and standardize company name"""
    if not company_name:
        return ""
    
    # Remove common suffixes for better matching
    suffixes_to_remove = [
        'LIMITED', 'LTD', 'PLC', 'LLC', 'INC', 'CORP', 'CORPORATION',
        'COMPANY', 'CO', 'LLP', 'LP', 'PARTNERSHIP'
    ]
    
    clean_name = company_name.upper().strip()
    
    for suffix in suffixes_to_remove:
        if clean_name.endswith(f' {suffix}'):
            clean_name = clean_name[:-len(suffix)-1].strip()
        elif clean_name.endswith(suffix):
            clean_name = clean_name[:-len(suffix)].strip()
    
    return clean_name

def extract_domain_from_company_name(company_name: str) -> str:
    """Extract or guess domain from company name"""
    if not company_name:
        return ""
    
    # Clean the company name
    clean_name = clean_company_name(company_name)
    
    # Remove non-alphanumeric characters and convert to lowercase
    domain_base = re.sub(r'[^a-zA-Z0-9\s]', '', clean_name).lower()
    
    # Remove common words
    common_words = ['the', 'and', 'of', 'for', 'in', 'on', 'at', 'to', 'by', 'with']
    words = domain_base.split()
    filtered_words = [word for word in words if word not in common_words]
    
    # Join words and create domain guess
    domain_guess = ''.join(filtered_words[:3])  # Limit to first 3 words
    
    return f"{domain_guess}.com" if domain_guess else ""

def validate_enrichment_data(data: Dict) -> Dict[str, Any]:
    """Validate enrichment data quality"""
    validation_result = {
        'is_valid': True,
        'quality_score': 0,
        'issues': [],
        'warnings': []
    }
    
    if not data:
        validation_result['is_valid'] = False
        validation_result['issues'].append("No data provided")
        return validation_result
    
    # Check required fields
    required_fields = ['name', 'domain']
    for field in required_fields:
        if not data.get(field):
            validation_result['issues'].append(f"Missing required field: {field}")
    
    # Validate specific fields
    if data.get('employee_count'):
        try:
            emp_count = int(data['employee_count'])
            if emp_count < 0:
                validation_result['warnings'].append("Negative employee count")
            elif emp_count > 1000000:
                validation_result['warnings'].append("Unusually high employee count")
        except (ValueError, TypeError):
            validation_result['warnings'].append("Invalid employee count format")
    
    # Validate domain format
    if data.get('domain'):
        domain = data['domain']
        if not re.match(r'^[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', domain):
            validation_result['warnings'].append("Invalid domain format")
    
    # Calculate quality score
    quality_fields = [
        'name', 'domain', 'industry', 'employee_count', 'annual_revenue',
        'description', 'founded_year', 'location', 'technologies', 'social_profiles'
    ]
    
    present_fields = sum(1 for field in quality_fields if data.get(field))
    validation_result['quality_score'] = (present_fields / len(quality_fields)) * 100
    
    # Overall validity
    validation_result['is_valid'] = len(validation_result['issues']) == 0
    
    return validation_result

def format_currency(amount: Any, currency: str = "GBP") -> str:
    """Format currency amounts"""
    if not amount:
        return "N/A"
    
    try:
        amount_float = float(amount)
        if currency == "GBP":
            return f"£{amount_float:,.0f}"
        elif currency == "USD":
            return f"${amount_float:,.0f}"
        else:
            return f"{amount_float:,.0f} {currency}"
    except (ValueError, TypeError):
        return str(amount)

def format_employee_count(count: Any) -> str:
    """Format employee count with ranges"""
    if not count:
        return "N/A"
    
    try:
        count_int = int(count)
        if count_int == 1:
            return "1 employee"
        elif count_int < 10:
            return f"{count_int} employees"
        elif count_int < 50:
            return f"{count_int} employees (Small)"
        elif count_int < 250:
            return f"{count_int} employees (Medium)"
        else:
            return f"{count_int:,} employees (Large)"
    except (ValueError, TypeError):
        return str(count)

def sanitize_filename(filename: str) -> str:
    """Sanitize filename for safe file operations"""
    # Remove or replace invalid characters
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        filename = filename.replace(char, '_')
    
    # Limit length
    if len(filename) > 200:
        filename = filename[:200]
    
    return filename.strip()

def normalize_borough_name(borough_name: str) -> str:
    """Normalize borough name to handle all variations consistently
    
    Handles cases like:
    - "Kensington & Chelsea" -> "Kensington and Chelsea"
    - "hammersmith and fulham" -> "Hammersmith and Fulham"
    - "Tower  Hamlets" -> "Tower Hamlets"
    
    Args:
        borough_name: Raw borough name from various sources
        
    Returns:
        Normalized borough name for consistent matching
    """
    if not borough_name:
        return ""
    
    # Basic cleaning
    normalized = borough_name.strip()
    
    # Replace & with 'and'
    normalized = re.sub(r'\s*&\s*', ' and ', normalized)
    
    # Clean up multiple whitespace
    normalized = re.sub(r'\s+', ' ', normalized)
    
    # Apply proper title case
    normalized = normalized.title()
    
    # Handle special cases that need manual correction
    special_cases = {
        'Kensington And Chelsea': 'Kensington and Chelsea',
        'Hammersmith And Fulham': 'Hammersmith and Fulham',
        'Barking And Dagenham': 'Barking and Dagenham',
        'Richmond Upon Thames': 'Richmond upon Thames',
        'Kingston Upon Thames': 'Kingston upon Thames'
    }
    
    # Apply special case corrections
    if normalized in special_cases:
        normalized = special_cases[normalized]
    
    return normalized

def is_outline(app: Dict[str, Any]) -> bool:
    """Centralized function to detect outline planning applications
    
    This function consolidates all outline detection logic into a single,
    reliable method to eliminate duplication and ensure consistency.
    
    Detection methods (in order of reliability):
    1. Application type field contains 'outline' 
    2. Reference code ends with 'OUT' (case-insensitive)
    3. Description must contain specific outline phrases
    
    Args:
        app: Planning application dict with fields like:
            - application_type: Official application type
            - lpa_app_no: Planning reference number
            - description: Application description
            - development_description: Alternative description field
            - proposal_description: Alternative description field
            - work_description: Alternative description field
    
    Returns:
        True if application is detected as outline, False otherwise
    """
    if not app or not isinstance(app, dict):
        return False
    
    # Method 1: Check application type field (most reliable)
    app_type = str(app.get('application_type', '')).lower().strip()
    if 'outline' in app_type:
        return True
    
    # Method 2: Check reference code patterns (very reliable)
    reference = str(app.get('lpa_app_no', '') or app.get('reference', '')).upper().strip()
    if reference.endswith('OUT') or reference.endswith('OUTL') or '/OUT' in reference:
        return True
    
    # Method 3: Check description fields for specific outline phrases (strict)
    description_fields = [
        app.get('description', ''),
        app.get('development_description', ''), 
        app.get('proposal_description', ''),
        app.get('work_description', '')
    ]
    
    # Combine all description text
    full_description = ' '.join(str(field) for field in description_fields if field).lower()
    
    # Strict outline phrase requirements
    required_phrases = [
        'outline planning application',
        'outline planning permission'
    ]
    
    # Must contain one of the required phrases
    for phrase in required_phrases:
        if phrase in full_description:
            return True
    
    
    return False

def create_outline_elasticsearch_query() -> Dict[str, Any]:
    """Create Elasticsearch query for server-side outline filtering
    
    Strict filtering that requires descriptions to contain specific
    outline phrases: "outline planning application" or "outline planning permission".
    Also includes reference number patterns with "OUT".
    
    Returns:
        Elasticsearch query dict filtering for outline applications
    """
    return {
        "bool": {
            "should": [
                # Reference number patterns (existing working queries)
                {"match": {"lpa_app_no": "OUT"}},
                {"query_string": {"default_field": "lpa_app_no", "query": "OUT"}},
                
                # Strict description field searches for specific phrases
                {"match_phrase": {"description": "outline planning application"}},
                {"match_phrase": {"description": "outline planning permission"}},
                {"match_phrase": {"development_description": "outline planning application"}},
                {"match_phrase": {"development_description": "outline planning permission"}},
                {"match_phrase": {"proposal_description": "outline planning application"}},
                {"match_phrase": {"proposal_description": "outline planning permission"}},
                {"match_phrase": {"work_description": "outline planning application"}},
                {"match_phrase": {"work_description": "outline planning permission"}}
            ],
            "minimum_should_match": 1
        }
    }

def generate_cache_key(search_params: Dict[str, Any]) -> str:
    """Generate a unique cache key that includes filter signature
    
    This ensures that different search filters create different cache entries,
    preventing stale data issues where cached results don't match current filters.
    
    Args:
        search_params: Dict containing search parameters like:
            - local_authority: Borough name
            - application_type: Type filter (including 'Outline')
            - start_date: Date filter
            - decision_status: Decision filter
            - limit: Result limit
            - enable_outline_filter: Whether outline server-side filtering is enabled
    
    Returns:
        Unique cache key string that includes all significant filter parameters
    """
    # Extract key parameters that affect results
    key_components = []
    
    # Add each parameter that affects search results
    if search_params.get('local_authority'):
        key_components.append(f"auth:{normalize_borough_name(search_params['local_authority'])}")
    
    if search_params.get('application_type'):
        key_components.append(f"type:{search_params['application_type']}")
    
    if search_params.get('start_date'):
        key_components.append(f"date:{search_params['start_date']}")
    
    if search_params.get('decision_status'):
        key_components.append(f"decision:{search_params['decision_status']}")
    
    if search_params.get('limit'):
        key_components.append(f"limit:{search_params['limit']}")
    
    # Important: Include outline filter flag
    if search_params.get('enable_outline_filter'):
        key_components.append("outline:server-filtered")
    
    # Create hash from components
    key_string = "|".join(sorted(key_components))
    
    # Use a hash for consistent length
    import hashlib
    return hashlib.md5(key_string.encode()).hexdigest()[:16]
import pandas as pd

def format_company_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Placeholder: clean/format company data.
    For now, just returns the DataFrame unchanged.
    """
    return df

def export_to_excel(df: pd.DataFrame, filename: str) -> str:
    """
    Export a DataFrame to an Excel file.
    """
    # Requires: openpyxl (already in your requirements)
    df.to_excel(filename, index=False)
    return filename

def validate_company_number(company_number: str) -> bool:
    """
    Basic validation for company numbers.
    Adjust logic to match your business rules.
    """
    return company_number.isdigit() and len(company_number) in (8, 10)
